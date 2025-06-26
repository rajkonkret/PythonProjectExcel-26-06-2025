from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws.append([1, 2, 3])

wb.save('sample.xlsx')
wb.close()
print("Excel file has been created succesfully")

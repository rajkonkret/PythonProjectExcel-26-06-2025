import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook('videogamesales2.xlsx')
# ws = wb.active


ws = wb['Video Games Sales Data']

ws['A1'].font = Font(bold=True, size=12)

for cell in ws[1:1]:
    cell.font = Font(bold=True, size=12)

wb.save("videogamesales2.xlsx")
wb.close()

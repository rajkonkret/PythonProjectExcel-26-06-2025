import openpyxl

import openpyxl

wb = openpyxl.load_workbook('videogamesales2.xlsx')
# ws = wb.active

ws = wb['vgsales']

ws['P1'] = 'Average Sales'
ws['P2'] = '= AVERAGE(K2:K16199)'

wb.save('videogamesales2.xlsx')
wb.close()

ws['Q1'] = "Number of Populated Cells"
ws['Q2'] = '=COUNTA(E2:E16220)'

wb.save('videogamesales2.xlsx')
wb.close()

ws['R1'] = "Number of rows with Sports Genre"
ws['R2'] = '=COUNTIF(E2:E16620, "sports")'

wb.save('videogamesales2.xlsx')
wb.close()

ws['S1'] = "Total sports Sales"
ws['S2'] = '=SUMIF(E2:E16220,"Sports", K2:K16220)'

wb.save('videogamesales2.xlsx')
wb.close()

ws['T1'] = "Rounded Sum of sports Sales"
ws['T2'] = '=CEILING(S2, 25)'

wb.save('videogamesales2.xlsx')
wb.close()
